import openpyxl
import json


class MyExcel:
    assignment_excel_path = "data/assignment.xlsx"
    quiz_excel_path = "data/quiz.xlsx"

    @staticmethod
    def getAssignmentTest():
        wb = openpyxl.load_workbook(MyExcel.assignment_excel_path)
        sheet = wb["Data"]
        js = {
            "assignment": []
        }
        url = None
        path = None
        for col in sheet.iter_cols():
            header = col[0].value
            if header == "URL":
                url = col
            elif header == "Filepath":
                path = col

        idx = 1
        while idx < len(url):
            test = {
                "url": url[idx].value,
                "filepath": path[idx].value
            }
            js["assignment"].append(test)
            idx += 1
        return json.dumps(js)

    @staticmethod
    def setAssignmentResult(idx, res):
        wb = openpyxl.load_workbook(MyExcel.assignment_excel_path)
        sheet = wb["Data"]
        resCol = None
        for col in sheet.iter_cols():
            header = col[0].value
            if header == "Result":
                resCol = col
        if idx >= len(res):
            print("Wrong idx!!")
            return
        resCol[idx].value = res
        wb.save(MyExcel.assignment_excel_path)

    @staticmethod
    def getQuizTest():
        wb = openpyxl.load_workbook(MyExcel.quiz_excel_path)
        sheet = wb["Data"]
        js = {
            "quiz": []
        }
        url = None
        answer = None
        for col in sheet.iter_cols():
            header = col[0].value
            if header == "URL":
                url = col
            elif header == "Answer":
                answer = col
        idx = 1
        while idx < len(url):
            test = {
                "url": url[idx].value,
                "answer": eval(answer[idx].value)
            }
            js["quiz"].append(test)
            idx += 1
        return json.dumps(js)

    @staticmethod
    def setQuizResult(idx, res):
        wb = openpyxl.load_workbook(MyExcel.quiz_excel_path)
        sheet = wb["Data"]
        resCol = None
        for col in sheet.iter_cols():
            header = col[0].value
            if header == "Result":
                resCol = col
        if idx >= len(res):
            print("Wrong idx!!")
            return
        resCol[idx].value = res
        wb.save(MyExcel.quiz_excel_path)


